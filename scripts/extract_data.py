#!/usr/bin/env python3
"""
Extract RBI macro indicators from both Excel workbooks into a single
consolidated JSON dataset (with CSV twin) ready for the HTML dashboard.

Source files (relative to the script directory):
  - 50 Macroeconomic Indicators.xlsx     (Weekly / Fortnightly / Monthly / Quarterly)
  - Other Macroeconomic Indicators.xlsx  (Daily   / Weekly   / Monthly / Quarterly)

Output: data/macro_data.json — single file the dashboard loads.
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path
from datetime import datetime, date

import openpyxl

# Local theme taxonomy
sys.path.insert(0, str(Path(__file__).resolve().parent))
from themes import classify, get_meta, THEME_ORDER

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
FOLDER_50 = BASE_DIR / "50 Macroeconomic Indicators.xlsx"
FOLDER_OTHER = BASE_DIR / "Other Macroeconomic Indicators.xlsx"
OUT_JSON = BASE_DIR / "data" / "macro_data.json"

SOURCE_URL = "Official Indian macroeconomic indicators, weekly + monthly"

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def normalise_header(s):
    if s is None:
        return ""
    s = str(s).replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def to_iso_date(v):
    """Convert date-like input to ISO date string or None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s or s.lower() in ("wh", "na", "n/a", "-", "—"):
            return None
        # 29-Jun-2026
        m = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{4})$", s)
        if m:
            d, mon, y = int(m.group(1)), m.group(2).lower(), int(m.group(3))
            if mon in MONTHS:
                try:
                    return date(y, MONTHS[mon], d).isoformat()
                except ValueError:
                    pass
        # May-2026 / Jun-2026
        m = re.match(r"^([A-Za-z]{3,9})-(\d{4})$", s)
        if m:
            mon, y = m.group(1).lower()[:3], int(m.group(2))
            if mon in MONTHS:
                return date(y, MONTHS[mon], 1).isoformat()
        # 2026-Q1
        m = re.match(r"^(\d{4})-Q([1-4])$", s)
        if m:
            y, q = int(m.group(1)), int(m.group(2))
            return date(y, (q - 1) * 3 + 1, 1).isoformat()
        return s  # leave as-is
    return v


def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "-", "—", "NA", "N/A", "wh"):
            return None
        try:
            if "." in s or "e" in s.lower():
                return float(s)
            return int(s)
        except ValueError:
            return v
    return v


def detect_header_row(rows):
    """Find the row that contains 'Period' or 'Reporting Date' as the period column."""
    for i, r in enumerate(rows[:8]):
        for c in r:
            if c and isinstance(c, str) and ("Period" in c or "Reporting Date" in c):
                return i
    # fall back: first row that has any non-None values
    for i, r in enumerate(rows[:5]):
        if any(c is not None for c in r):
            return i
    return 0


def _looks_like_units_row(row):
    """A unit row has lots of short string cells like 'Index', 'Per cent', '₹ Crore'."""
    strs = [c for c in row if isinstance(c, str) and c.strip()]
    if len(strs) < 2:
        return False
    if not all(len(s) < 30 for s in strs):
        return False
    # None of the cells should be a date
    if any(isinstance(c, (date, datetime)) for c in row):
        return False
    return True


def detect_units_row(rows, header_idx):
    """The unit row sits immediately before *or* after the header row."""
    # Try after first (some workbooks put units row 3 when headers are row 2)
    if header_idx + 1 < len(rows) and _looks_like_units_row(rows[header_idx + 1]):
        return header_idx + 1
    # Then try before
    if header_idx > 0 and _looks_like_units_row(rows[header_idx - 1]):
        return header_idx - 1
    return None


def detect_title_row(rows):
    for r in rows[:5]:
        for c in r:
            if c and isinstance(c, str) and "Macro" in c:
                return c.strip()
    return ""


def extract_sheet(ws, file_label, sheet_label, frequency):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return None

    title = detect_title_row(rows)
    header_idx = detect_header_row(rows)
    units_idx = detect_units_row(rows, header_idx)

    # Headers live in row header_idx; trim leading None columns
    raw_headers = list(rows[header_idx])
    first_data_col = 0
    for i, h in enumerate(raw_headers):
        if h is not None and (not isinstance(h, str) or normalise_header(h) != ""):
            first_data_col = i
            break

    headers = [normalise_header(h) if h else "" for h in raw_headers[first_data_col:]]
    if headers and not headers[0]:
        headers = headers[1:]

    # Units (if any) — match by column index from the same offset
    units = []
    if units_idx is not None:
        urow = list(rows[units_idx])[first_data_col:]
        for i in range(len(headers)):
            u = urow[i] if i < len(urow) else None
            units.append(normalise_header(u) if isinstance(u, str) else "")

    # Data rows
    series = [{"name": h, "unit": (units[i] if i < len(units) else ""),
               "values": []} for i, h in enumerate(headers)]

    n = len(headers)
    raw_rows = []
    for r in rows[header_idx + 1:]:
        if all(c is None for c in r):
            continue
        cells = list(r[first_data_col:])
        # skip footer/blank rows (period must be set)
        if not cells or cells[0] is None or (isinstance(cells[0], str) and not cells[0].strip()):
            continue
        # skip the 'wh' weekend/holiday rows
        if isinstance(cells[0], str) and cells[0].strip().lower() == "wh":
            continue

        while len(cells) < n:
            cells.append(None)
        cells = cells[:n]

        normalised = []
        for i, v in enumerate(cells):
            if i == 0:
                normalised.append(to_iso_date(v))
            else:
                normalised.append(to_number(v))
        raw_rows.append(normalised)

    # ---- Sort ASCENDING by period so values[-1] = latest, values[0] = earliest ----
    def sort_key(row):
        p = row[0]
        if p is None:
            return ("",)  # empties last
        if isinstance(p, str):
            return (p,)
        return (p,)
    raw_rows.sort(key=sort_key)

    for row in raw_rows:
        for i, v in enumerate(row):
            series[i]["values"].append(v)

    if not series:
        return None

    # ---- Compute YoY and MoM (period-over-period) derived series ----
    # Use date-based lookback where possible so the math is robust to sparse data.
    PERIODS_PER_YEAR = {
        "daily": 365, "weekly": 52, "fortnightly": 26, "monthly": 12, "quarterly": 4,
    }
    lookback = PERIODS_PER_YEAR.get(frequency, 12)
    periods = series[0]["values"]

    def date_obj(s):
        if not isinstance(s, str): return None
        try: return date.fromisoformat(s[:10])
        except: return None

    for i in range(1, len(series)):
        vals = series[i]["values"]
        yoy = [None] * len(vals)
        mom = [None] * len(vals)
        for j, v in enumerate(vals):
            if v is None or not isinstance(v, (int, float)) or v == 0:
                continue
            # Find the index that is ~1 year before j
            if j < lookback:
                # fallback: simple index-based lookback
                yoy_idx = j - lookback
                if yoy_idx >= 0 and isinstance(vals[yoy_idx], (int, float)) and vals[yoy_idx] != 0:
                    yoy[j] = ((v - vals[yoy_idx]) / abs(vals[yoy_idx])) * 100
            else:
                target_d = date_obj(periods[j])
                if target_d:
                    best, best_diff = None, None
                    for k in range(j - 1, max(-1, j - lookback * 2), -1):
                        d = date_obj(periods[k])
                        if d is None: continue
                        diff = abs((target_d - d).days - 365)
                        if best_diff is None or diff < best_diff:
                            best_diff = diff; best = k
                            if diff < 2: break
                    if best is not None and isinstance(vals[best], (int, float)) and vals[best] != 0 and best_diff is not None and best_diff < 15:
                        yoy[j] = ((v - vals[best]) / abs(vals[best])) * 100
            # MoM/PoP: previous period (j-1)
            if j > 0 and isinstance(vals[j-1], (int, float)) and vals[j-1] != 0:
                mom[j] = ((v - vals[j-1]) / abs(vals[j-1])) * 100
        series[i]["derived"] = {"yoy": yoy, "mom": mom}

    # ---- Attach theme metadata to each series ----
    for ser in series:
        ser["theme"] = classify(ser["name"])

    # Sheet-level: list of themes present (for filter chips etc.)
    present = sorted({s["theme"] for s in series[1:]}, key=lambda t: THEME_ORDER.index(t) if t in THEME_ORDER else 999)

    sid = (
        f"{file_label}_{sheet_label.lower()}"
        .replace(" ", "_").replace("-", "_")
    )

    return {
        "id": sid,
        "file": file_label,
        "name": sheet_label,
        "frequency": frequency,
        "title": title or f"{file_label} — {sheet_label}",
        "rowCount": len(series[0]["values"]) if series else 0,
        "themes": present,
        "series": series,
    }


def extract_workbook(path, file_label, frequency_map):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = []
    for sn in wb.sheetnames:
        freq = frequency_map.get(sn, sn.lower())
        sheet = extract_sheet(wb[sn], file_label, sn, freq)
        if sheet and sheet["series"]:
            sheets.append(sheet)
    return sheets


def main():
    sheets = []
    sheets += extract_workbook(
        FOLDER_50,
        "50_macroeconomic_indicators",
        {
            "Weekly": "weekly",
            "Fortnightly": "fortnightly",
            "Monthly": "monthly",
            "Quarterly": "quarterly",
        },
    )
    sheets += extract_workbook(
        FOLDER_OTHER,
        "other_macroeconomic_indicators",
        {
            "Daily": "daily",
            "Weekly": "weekly",
            "Monthly": "monthly",
            "Quarterly": "quarterly",
        },
    )

    payload = {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "sourceUrl": SOURCE_URL,
        "files": {
            "50_macroeconomic_indicators": {
                "path": FOLDER_50.name,
                "source": "50 Macroeconomic Indicators",
            },
            "other_macroeconomic_indicators": {
                "path": FOLDER_OTHER.name,
                "source": "Other Macroeconomic Indicators",
            },
        },
        "themes": [
            {"id": tid, "label": get_meta(tid)["label"],
             "icon": get_meta(tid)["icon"], "color": get_meta(tid)["color"]}
            for tid in THEME_ORDER
        ],
        "sheets": sheets,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    total_series = sum(len(s["series"]) for s in sheets)
    total_points = sum(
        sum(1 for v in s["series"][i]["values"] if v is not None)
        for s in sheets
        for i in range(1, len(s["series"]))
    )
    print(f"Wrote {OUT_JSON}")
    print(f"  Sheets: {len(sheets)}")
    print(f"  Series: {total_series}")
    print(f"  Non-null data points: {total_points:,}")
    for s in sheets:
        cols = ", ".join(ser["name"][:32] for ser in s["series"][:6])
        print(f"  • {s['id']} ({s['rowCount']} rows): {cols}{' ...' if len(s['series'])>6 else ''}")


if __name__ == "__main__":
    main()